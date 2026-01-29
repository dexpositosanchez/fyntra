from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from typing import List, Optional
from datetime import datetime, date
from app.database import get_db
from app.models.comunidad import Comunidad
from app.models.inmueble import Inmueble
from app.models.incidencia import Incidencia
from app.models.actuacion import Actuacion
from app.models.proveedor import Proveedor
from app.models.usuario import Usuario
from app.api.dependencies import get_current_user
import io
import csv
from decimal import Decimal

router = APIRouter(prefix="/informes", tags=["informes"])

ROLES_ADMIN_INFORMES = ["super_admin", "admin_fincas", "admin_transportes"]


def verificar_admin_informes(usuario: Usuario):
    """Solo super_admin y administradores pueden consultar informes."""
    if usuario.rol not in ROLES_ADMIN_INFORMES:
        raise HTTPException(
            status_code=403,
            detail="Solo super administradores y administradores pueden consultar informes"
        )


def formatear_fecha_espanol(fecha_str: str) -> str:
    """Convierte fecha de formato YYYY-MM-DD a dd/mm/YYYY"""
    try:
        fecha = datetime.fromisoformat(fecha_str).date()
        return f"{fecha.day:02d}/{fecha.month:02d}/{fecha.year}"
    except:
        return fecha_str

def _build_pdf(titulo: str, subtitulo: str, headers: list, rows: list) -> bytes:
    """
    Genera un PDF simple en memoria usando ReportLab.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab no está instalado. Instálelo con: pip install reportlab")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()

    story = []
    story.append(Paragraph(titulo, styles["Title"]))
    story.append(Paragraph(subtitulo, styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f4f7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#2f343d")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e0e0e0")),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    story.append(table)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf

def get_first_and_last_day_of_month(year: int, month: int):
    """Obtiene el primer y último día del mes"""
    from calendar import monthrange
    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])
    return first_day, last_day

@router.get("/comunidades")
async def obtener_informes_comunidades(
    fecha_inicio: Optional[str] = Query(None, description="Fecha de inicio (YYYY-MM-DD)"),
    fecha_fin: Optional[str] = Query(None, description="Fecha de fin (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene informes de comunidades con sus inmuebles, número de incidencias y costes totales.
    Por defecto usa el mes actual si no se especifican fechas.
    """
    # Si no se especifican fechas, usar el mes actual
    if not fecha_inicio or not fecha_fin:
        today = date.today()
        fecha_inicio, fecha_fin = get_first_and_last_day_of_month(today.year, today.month)
        fecha_inicio_str = fecha_inicio.isoformat()
        fecha_fin_str = fecha_fin.isoformat()
    else:
        fecha_inicio_str = fecha_inicio
        fecha_fin_str = fecha_fin
    
    # Convertir a datetime para la consulta
    fecha_inicio_dt = datetime.fromisoformat(fecha_inicio_str).replace(hour=0, minute=0, second=0)
    fecha_fin_dt = datetime.fromisoformat(fecha_fin_str).replace(hour=23, minute=59, second=59)
    
    # Obtener todas las comunidades con sus inmuebles
    comunidades = db.query(Comunidad).all()
    
    resultado = []
    
    for comunidad in comunidades:
        # Obtener inmuebles de la comunidad
        inmuebles = db.query(Inmueble).filter(Inmueble.comunidad_id == comunidad.id).all()
        
        inmuebles_data = []
        total_incidencias_comunidad = 0
        total_coste_comunidad = Decimal('0.00')
        
        for inmueble in inmuebles:
            # Obtener incidencias del inmueble en el rango de fechas
            incidencias = db.query(Incidencia).filter(
                and_(
                    Incidencia.inmueble_id == inmueble.id,
                    Incidencia.fecha_alta >= fecha_inicio_dt,
                    Incidencia.fecha_alta <= fecha_fin_dt
                )
            ).all()
            
            # Calcular coste total de las actuaciones de estas incidencias y agrupar por estado
            coste_total = Decimal('0.00')
            incidencias_por_estado = {}
            
            for incidencia in incidencias:
                estado = incidencia.estado.value if hasattr(incidencia.estado, 'value') else str(incidencia.estado)
                if estado not in incidencias_por_estado:
                    incidencias_por_estado[estado] = {"count": 0, "coste": Decimal('0.00')}
                
                incidencias_por_estado[estado]["count"] += 1
                
                actuaciones = db.query(Actuacion).filter(
                    Actuacion.incidencia_id == incidencia.id
                ).all()
                for actuacion in actuaciones:
                    if actuacion.coste:
                        coste_actuacion = Decimal(str(actuacion.coste))
                        coste_total += coste_actuacion
                        incidencias_por_estado[estado]["coste"] += coste_actuacion
            
            num_incidencias = len(incidencias)
            total_incidencias_comunidad += num_incidencias
            total_coste_comunidad += coste_total
            
            inmuebles_data.append({
                "id": inmueble.id,
                "referencia": inmueble.referencia,
                "direccion": inmueble.direccion,
                "metros": float(inmueble.metros) if inmueble.metros else None,
                "tipo": inmueble.tipo,
                "num_incidencias": num_incidencias,
                "coste_total": float(coste_total),
                "incidencias_por_estado": {k: {"count": v["count"], "coste": float(v["coste"])} for k, v in incidencias_por_estado.items()}
            })
        
        # Calcular incidencias por estado a nivel de comunidad
        incidencias_comunidad_por_estado = {}
        for inmueble_data in inmuebles_data:
            if "incidencias_por_estado" in inmueble_data:
                for estado, datos in inmueble_data["incidencias_por_estado"].items():
                    if estado not in incidencias_comunidad_por_estado:
                        incidencias_comunidad_por_estado[estado] = {"count": 0, "coste": 0.0}
                    incidencias_comunidad_por_estado[estado]["count"] += datos["count"]
                    incidencias_comunidad_por_estado[estado]["coste"] += datos["coste"]
        
        # Solo agregar comunidades que tengan incidencias
        if total_incidencias_comunidad > 0:
            resultado.append({
                "id": comunidad.id,
                "nombre": comunidad.nombre,
                "cif": comunidad.cif,
                "direccion": comunidad.direccion,
                "num_incidencias": total_incidencias_comunidad,
                "coste_total": float(total_coste_comunidad),
                "incidencias_por_estado": incidencias_comunidad_por_estado,
                "inmuebles": [inm for inm in inmuebles_data if inm["num_incidencias"] > 0]  # Solo inmuebles con incidencias
            })
    
    return {
        "fecha_inicio": fecha_inicio_str,
        "fecha_fin": fecha_fin_str,
        "comunidades": resultado
    }

@router.get("/proveedores")
async def obtener_informes_proveedores(
    fecha_inicio: Optional[str] = Query(None, description="Fecha de inicio (YYYY-MM-DD)"),
    fecha_fin: Optional[str] = Query(None, description="Fecha de fin (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtiene informes de proveedores con número de incidencias y costes totales.
    Por defecto usa el mes actual si no se especifican fechas.
    """
    verificar_admin_informes(current_user)
    # Si no se especifican fechas, usar el mes actual
    if not fecha_inicio or not fecha_fin:
        today = date.today()
        fecha_inicio, fecha_fin = get_first_and_last_day_of_month(today.year, today.month)
        fecha_inicio_str = fecha_inicio.isoformat()
        fecha_fin_str = fecha_fin.isoformat()
    else:
        fecha_inicio_str = fecha_inicio
        fecha_fin_str = fecha_fin

    # Convertir a datetime para la consulta
    fecha_inicio_dt = datetime.fromisoformat(fecha_inicio_str).replace(hour=0, minute=0, second=0)
    fecha_fin_dt = datetime.fromisoformat(fecha_fin_str).replace(hour=23, minute=59, second=59)

    # Obtener todos los proveedores
    proveedores = db.query(Proveedor).all()
    
    resultado = []
    
    for proveedor in proveedores:
        # Obtener incidencias asignadas al proveedor en el rango de fechas
        incidencias = db.query(Incidencia).filter(
            and_(
                Incidencia.proveedor_id == proveedor.id,
                Incidencia.fecha_alta >= fecha_inicio_dt,
                Incidencia.fecha_alta <= fecha_fin_dt
            )
        ).all()
        
        # Calcular coste total de las actuaciones de estas incidencias y agrupar por estado
        coste_total = Decimal('0.00')
        incidencias_por_estado = {}
        
        for incidencia in incidencias:
            estado = incidencia.estado.value if hasattr(incidencia.estado, 'value') else str(incidencia.estado)
            if estado not in incidencias_por_estado:
                incidencias_por_estado[estado] = {"count": 0, "coste": Decimal('0.00')}
            
            incidencias_por_estado[estado]["count"] += 1
            
            actuaciones = db.query(Actuacion).filter(
                Actuacion.incidencia_id == incidencia.id
            ).all()
            for actuacion in actuaciones:
                if actuacion.coste:
                    coste_actuacion = Decimal(str(actuacion.coste))
                    coste_total += coste_actuacion
                    incidencias_por_estado[estado]["coste"] += coste_actuacion
        
        # Solo agregar proveedores que tengan incidencias
        if len(incidencias) > 0:
            resultado.append({
                "id": proveedor.id,
                "nombre": proveedor.nombre,
                "email": proveedor.email,
                "telefono": proveedor.telefono,
                "especialidad": proveedor.especialidad,
                "num_incidencias": len(incidencias),
                "coste_total": float(coste_total),
                "incidencias_por_estado": {k: {"count": v["count"], "coste": float(v["coste"])} for k, v in incidencias_por_estado.items()}
            })
    
    return {
        "fecha_inicio": fecha_inicio_str,
        "fecha_fin": fecha_fin_str,
        "proveedores": resultado
    }

@router.get("/comunidades/exportar/{formato}")
async def exportar_informes_comunidades(
    formato: str,
    fecha_inicio: Optional[str] = Query(None, description="Fecha de inicio (YYYY-MM-DD)"),
    fecha_fin: Optional[str] = Query(None, description="Fecha de fin (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Exporta informes de comunidades en formato PDF, Excel o CSV.
    """
    # Obtener los datos
    datos = await obtener_informes_comunidades(fecha_inicio, fecha_fin, db, current_user)
    
    if formato.lower() == "csv":
        # Generar CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Período en formato español
        fecha_inicio_es = formatear_fecha_espanol(datos['fecha_inicio'])
        fecha_fin_es = formatear_fecha_espanol(datos['fecha_fin'])
        writer.writerow([f"Período: {fecha_inicio_es} - {fecha_fin_es}"])
        writer.writerow([])  # Línea en blanco
        
        # Encabezados
        writer.writerow(["Comunidad", "CIF", "Dirección", "Inmueble", "Referencia Inmueble", "Nº Incidencias", "Coste Total (€)"])
        
        # Datos
        for comunidad in datos["comunidades"]:
            # Fila de comunidad (sin inmueble específico)
            writer.writerow([
                comunidad["nombre"],
                comunidad["cif"] or "",
                comunidad["direccion"] or "",
                "",
                "",
                comunidad["num_incidencias"],
                f"{comunidad['coste_total']:.2f}"
            ])
            
            # Filas de inmuebles
            for inmueble in comunidad["inmuebles"]:
                writer.writerow([
                    "",
                    "",
                    "",
                    inmueble["direccion"] or "",
                    inmueble["referencia"],
                    inmueble["num_incidencias"],
                    f"{inmueble['coste_total']:.2f}"
                ])
        
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=informes_comunidades_{datos['fecha_inicio']}_{datos['fecha_fin']}.csv"
            }
        )
    
    elif formato.lower() == "excel":
        # Para Excel necesitaríamos openpyxl o xlsxwriter
        # Por ahora devolvemos CSV con extensión .xlsx (o mejor, implementamos con openpyxl)
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Informes Comunidades"
            
            # Período en formato español
            fecha_inicio_es = formatear_fecha_espanol(datos['fecha_inicio'])
            fecha_fin_es = formatear_fecha_espanol(datos['fecha_fin'])
            periodo_cell = ws.cell(row=1, column=1, value=f"Período: {fecha_inicio_es} - {fecha_fin_es}")
            periodo_cell.font = Font(bold=True, size=12)
            ws.append([])  # Línea en blanco
            
            # Encabezados
            headers = ["Comunidad", "CIF", "Dirección", "Inmueble", "Referencia Inmueble", "Nº Incidencias", "Coste Total (€)"]
            ws.append(headers)
            
            # Estilo para encabezados
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            for comunidad in datos["comunidades"]:
                # Fila de comunidad
                ws.append([
                    comunidad["nombre"],
                    comunidad["cif"] or "",
                    comunidad["direccion"] or "",
                    "",
                    "",
                    comunidad["num_incidencias"],
                    comunidad["coste_total"]
                ])
                
                # Filas de inmuebles
                for inmueble in comunidad["inmuebles"]:
                    ws.append([
                        "",
                        "",
                        "",
                        inmueble["direccion"] or "",
                        inmueble["referencia"],
                        inmueble["num_incidencias"],
                        inmueble["coste_total"]
                    ])
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.read(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=informes_comunidades_{datos['fecha_inicio']}_{datos['fecha_fin']}.xlsx"
                }
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl no está instalado. Instálelo con: pip install openpyxl")
    
    elif formato.lower() == "pdf":
        titulo = "Informe de costes por comunidades"
        fecha_inicio_es = formatear_fecha_espanol(datos['fecha_inicio'])
        fecha_fin_es = formatear_fecha_espanol(datos['fecha_fin'])
        subtitulo = f"Período: {fecha_inicio_es} - {fecha_fin_es}"
        headers = ["Comunidad", "CIF", "Dirección", "Inmueble", "Referencia", "Nº Incidencias", "Coste Total (€)"]
        rows = []
        for comunidad in datos["comunidades"]:
            rows.append([
                comunidad["nombre"],
                comunidad.get("cif") or "",
                comunidad.get("direccion") or "",
                "",
                "",
                str(comunidad["num_incidencias"]),
                f"{comunidad['coste_total']:.2f}",
            ])
            for inmueble in comunidad.get("inmuebles", []):
                rows.append([
                    "",
                    "",
                    "",
                    inmueble.get("direccion") or "",
                    inmueble.get("referencia") or "",
                    str(inmueble["num_incidencias"]),
                    f"{inmueble['coste_total']:.2f}",
                ])

        pdf_bytes = _build_pdf(titulo, subtitulo, headers, rows)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=informes_comunidades_{datos['fecha_inicio']}_{datos['fecha_fin']}.pdf"
            },
        )
    
    else:
        raise HTTPException(status_code=400, detail="Formato no válido. Use: csv, excel o pdf")

@router.get("/proveedores/exportar/{formato}")
async def exportar_informes_proveedores(
    formato: str,
    fecha_inicio: Optional[str] = Query(None, description="Fecha de inicio (YYYY-MM-DD)"),
    fecha_fin: Optional[str] = Query(None, description="Fecha de fin (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Exporta informes de proveedores en formato PDF, Excel o CSV.
    """
    # Obtener los datos
    datos = await obtener_informes_proveedores(fecha_inicio, fecha_fin, db, current_user)
    
    if formato.lower() == "csv":
        # Generar CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Período en formato español
        fecha_inicio_es = formatear_fecha_espanol(datos['fecha_inicio'])
        fecha_fin_es = formatear_fecha_espanol(datos['fecha_fin'])
        writer.writerow([f"Período: {fecha_inicio_es} - {fecha_fin_es}"])
        writer.writerow([])  # Línea en blanco
        
        # Encabezados
        writer.writerow(["Proveedor", "Email", "Teléfono", "Especialidad", "Nº Incidencias", "Coste Total (€)"])
        
        # Datos
        for proveedor in datos["proveedores"]:
            writer.writerow([
                proveedor["nombre"],
                proveedor["email"] or "",
                proveedor["telefono"] or "",
                proveedor["especialidad"] or "",
                proveedor["num_incidencias"],
                f"{proveedor['coste_total']:.2f}"
            ])
        
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=informes_proveedores_{datos['fecha_inicio']}_{datos['fecha_fin']}.csv"
            }
        )
    
    elif formato.lower() == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Informes Proveedores"
            
            # Encabezados
            headers = ["Proveedor", "Email", "Teléfono", "Especialidad", "Nº Incidencias", "Coste Total (€)"]
            ws.append(headers)
            
            # Estilo para encabezados
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            for proveedor in datos["proveedores"]:
                ws.append([
                    proveedor["nombre"],
                    proveedor["email"] or "",
                    proveedor["telefono"] or "",
                    proveedor["especialidad"] or "",
                    proveedor["num_incidencias"],
                    proveedor["coste_total"]
                ])
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.read(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=informes_proveedores_{datos['fecha_inicio']}_{datos['fecha_fin']}.xlsx"
                }
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl no está instalado. Instálelo con: pip install openpyxl")
    
    elif formato.lower() == "pdf":
        titulo = "Informe de costes por proveedores"
        fecha_inicio_es = formatear_fecha_espanol(datos['fecha_inicio'])
        fecha_fin_es = formatear_fecha_espanol(datos['fecha_fin'])
        subtitulo = f"Período: {fecha_inicio_es} - {fecha_fin_es}"
        headers = ["Proveedor", "Email", "Teléfono", "Especialidad", "Nº Incidencias", "Coste Total (€)"]
        rows = []
        for proveedor in datos["proveedores"]:
            rows.append([
                proveedor["nombre"],
                proveedor.get("email") or "",
                proveedor.get("telefono") or "",
                proveedor.get("especialidad") or "",
                str(proveedor["num_incidencias"]),
                f"{proveedor['coste_total']:.2f}",
            ])

        pdf_bytes = _build_pdf(titulo, subtitulo, headers, rows)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=informes_proveedores_{datos['fecha_inicio']}_{datos['fecha_fin']}.pdf"
            },
        )
    
    else:
        raise HTTPException(status_code=400, detail="Formato no válido. Use: csv, excel o pdf")
